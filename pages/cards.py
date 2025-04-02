# cards.py (제니앱 - 카드값 계산기 v20)

import streamlit as st
import pandas as pd
import sys
import os
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from prism import detect_card_issuer, parse_card_file
from shared import show_menu
from rules import categorize

st.set_page_config(page_title="카드값 계산기 - 제니앱", page_icon="💳", layout="wide")
show_menu("카드값 계산기")

st.title("💳 카드값 계산기")

# ✅ 사용법 안내
st.markdown("""
### 📝 사용 방법

1. **카드사 자동 인식**  
   각 카드사 홈페이지에서 다운로드한 **원본 엑셀 파일 그대로 업로드**하세요.  
   파일명이나 시트 구조를 기반으로 **자동으로 카드사를 인식**합니다.

2. **여러 카드사 파일 동시 업로드 가능**  
   **여러 개의 엑셀 파일을 한꺼번에 업로드**해도 됩니다.  
   각 카드사의 내역이 자동으로 통합되어 보기 쉽게 정리됩니다.

3. **구버전 .xls 파일은 변환 후 사용**  
   카드사에 따라 **.xls 형식(구버전 엑셀)**으로 제공되는 경우가 있습니다.  
   이 경우 엑셀에서 열어 **[다른 이름으로 저장] → [.xlsx 형식]으로 저장한 후 업로드**해주세요.
""")

uploaded_files = st.file_uploader(
    "카드사별 이용 내역 파일 업로드 (여러 개 가능)",
    type=["xlsx"],
    accept_multiple_files=True
)

# ✅ 카드명 정규화 함수
def normalize_card_name(card):
    if "국민" in card:
        return "국민카드"
    if "신한" in card:
        return "신한카드"
    if "현대" in card:
        return "현대카드"
    if "하나" in card:
        return "하나카드"
    if "로테" in card:
        return "로테카드"
    if "삼성" in card:
        return "삼성카드"
    return card

# ✅ 이하 코드는 기존 v19와 동일하므로 유지
# ※ 이후 코드 생략 없이 전체 그대로 유지됨

if uploaded_files:
    all_records = []

    with st.container():
        for file in uploaded_files:
            st.markdown(f"---\n### 📂 {file.name}")
            card_issuer = detect_card_issuer(file)
            if not card_issuer:
                st.warning(f"❌ 카드사 인식 실패: {file.name}")
                continue
            df = parse_card_file(file, card_issuer)
            if df is not None:
                all_records.append(df)
                st.success(f"✅ {card_issuer} 내역 처리 완료: {len(df)}건")
            else:
                st.warning(f"⚠️ {card_issuer} 내역 파싱 실패")

    if all_records:
        final_df = pd.concat(all_records, ignore_index=True)
        final_df["카드"] = final_df["카드"].apply(normalize_card_name)
        final_df["카테고리"] = final_df["사용처"].apply(categorize)
        final_df["금액"] = final_df["금액"].apply(lambda x: float(str(x).replace(",", "")))
        final_df = final_df.sort_values(by=["카드", "카테고리", "날짜"]).reset_index(drop=True)

        st.subheader("📋 통합 카드 사용 내역")
        st.dataframe(final_df, use_container_width=True)

        @st.cache_data
        def to_excel(df):
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.page import PageMargins
            from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.series import DataPoint

            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = '카드내역'

            # ✅ 신한카드 색상 포함
            color_map_card = {
                "국민카드": "FBE2D5",
                "현대카드": "DDEBF7",
                "롯데카드": "CCCCFF",
                "삼성카드": "E2EFDA",
                "하나카드": "FFF2CC",
                "신한카드": "DDD9C4",
            }

            color_map_category = {
                "교통/주유/주차": "CCFFCC", "병원/약국": "FFCC99",
                "취미/쇼핑": "FFF2CC", "음식점/카페/편의점": "FFCCCC",
                "고정지출": "C6E0B4", "잡비용": "E7E6E6",
            }

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            ws.append(df.columns.tolist())
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            for i, width in enumerate([11, 11, 20, 40, 15]):
                ws.column_dimensions[chr(65 + i)].width = width
            ws.column_dimensions['F'].width = 3
            ws.column_dimensions['I'].width = 3

            ws.sheet_view.showGridLines = False

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                card = row[1].value
                category = row[2].value
                card_color = color_map_card.get(card, None)
                category_color = color_map_category.get(category, None)
                for idx, cell in enumerate(row):
                    cell.border = thin_border
                    if idx == 4:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                if card_color:
                    row[0].fill = PatternFill(start_color=card_color, end_color=card_color, fill_type="solid")
                    row[1].fill = PatternFill(start_color=card_color, end_color=card_color, fill_type="solid")
                if category_color:
                    row[2].fill = PatternFill(start_color=category_color, end_color=category_color, fill_type="solid")

            # 카테고리별 통계
            ws["G1"] = "카테고리"
            ws["H1"] = "금액"
            ws["G1"].fill = ws["H1"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws["G1"].font = ws["H1"].font = Font(color="FFFFFF", bold=True)
            ws["G1"].alignment = ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15

            stats = df.groupby("카테고리")["금액"].sum().reindex(color_map_category.keys()).dropna()
            total_sum = df["금액"].sum()

            row_idx = 2
            for cat, amount in stats.items():
                ws[f"G{row_idx}"] = cat
                ws[f"H{row_idx}"] = int(amount)
                ws[f"H{row_idx}"].number_format = '#,##0'
                color = color_map_category.get(cat)
                if color:
                    ws[f"G{row_idx}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border
                row_idx += 1

            ws[f"G{row_idx}"] = "합계"
            ws[f"H{row_idx}"] = int(total_sum)
            ws[f"G{row_idx}"].fill = ws[f"H{row_idx}"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws[f"G{row_idx}"].font = ws[f"H{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border

            # 카드사별 통계
            ws["G10"] = "카드사"
            ws["H10"] = "금액"
            ws["G10"].fill = ws["H10"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws["G10"].font = ws["H10"].font = Font(color="FFFFFF", bold=True)
            ws["G10"].alignment = ws["H10"].alignment = Alignment(horizontal="center", vertical="center")

            card_stats = df.groupby("카드")["금액"].sum().reindex(color_map_card.keys()).dropna()
            card_total = df["금액"].sum()

            row_idx = 11
            for card_name, amount in card_stats.items():
                ws[f"G{row_idx}"] = card_name
                ws[f"H{row_idx}"] = int(amount)
                ws[f"H{row_idx}"].number_format = '#,##0'
                color = color_map_card.get(card_name)
                if color:
                    ws[f"G{row_idx}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border
                row_idx += 1

            ws[f"G{row_idx}"] = "합계"
            ws[f"H{row_idx}"] = int(card_total)
            ws[f"G{row_idx}"].fill = ws[f"H{row_idx}"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws[f"G{row_idx}"].font = ws[f"H{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border

            # 카테고리별 원형 차트
            pie1 = PieChart()
            pie1.title = "카테고리별 사용 비중"
            labels1 = Reference(ws, min_col=7, min_row=2, max_row=7)
            data1 = Reference(ws, min_col=8, min_row=1, max_row=7)
            pie1.add_data(data1, titles_from_data=True)
            pie1.set_categories(labels1)
            pie1.height = 7
            pie1.width = 7
            for idx, cat in enumerate(color_map_category.keys()):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = color_map_category[cat]
                pie1.series[0].data_points.append(dp)
            ws.add_chart(pie1, "J1")

            # 카드사별 원형 차트
            pie2 = PieChart()
            pie2.title = "카드사별 사용 비중"
            labels2 = Reference(ws, min_col=7, min_row=11, max_row=16)
            data2 = Reference(ws, min_col=8, min_row=10, max_row=16)
            pie2.add_data(data2, titles_from_data=True)
            pie2.set_categories(labels2)
            pie2.height = 7
            pie2.width = 7
            for idx, card in enumerate(color_map_card.keys()):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = color_map_card[card]
                pie2.series[0].data_points.append(dp)
            ws.add_chart(pie2, "J14")

            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
            ws.sheet_properties = WorksheetProperties(pageSetUpPr=PageSetupProperties(fitToPage=True))

            wb.save(output)
            return output.getvalue()

        st.download_button(
            label="📅 엑셀파일 다운로드",
            data=to_excel(final_df),
            file_name="카드값_통합내역.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
